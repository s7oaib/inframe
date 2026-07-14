"""
Inframe — Export Route

Generate PDF and Excel attendance reports with date range filtering.
"""

import io
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db, Attendance, ClassSession, Person

router = APIRouter(prefix="/api/v1/reports", tags=["reports"])


@router.get("/export")
def export_report(
    format: str = Query("pdf", description="pdf or excel"),
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    db: Session = Depends(get_db),
):
    """Export attendance report as PDF or Excel."""
    # Parse dates
    try:
        start = date.fromisoformat(from_date) if from_date else date.today().replace(day=1)
        end = date.fromisoformat(to_date) if to_date else date.today()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    # Query attendance data
    records = (
        db.query(Attendance, Person.name, ClassSession.date, ClassSession.name.label("session_name"))
        .outerjoin(Person, Attendance.usn == Person.usn)
        .join(ClassSession, Attendance.session_id == ClassSession.id)
        .filter(ClassSession.date >= start, ClassSession.date <= end)
        .order_by(ClassSession.date, Attendance.usn)
        .all()
    )

    if format.lower() == "excel":
        return _generate_excel(records, start, end)
    else:
        return _generate_pdf(records, start, end)


def _generate_pdf(records, start, end):
    """Generate a professional PDF attendance report."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=20,
        spaceAfter=6,
        textColor=colors.HexColor('#11161C'),
    )
    elements.append(Paragraph("Inframe Attendance Report", title_style))

    # Date range
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#4A535C'),
        spaceAfter=20,
        alignment=TA_CENTER,
    )
    elements.append(Paragraph(f"Period: {start.strftime('%d %b %Y')} — {end.strftime('%d %b %Y')}", subtitle_style))
    elements.append(Spacer(1, 10))

    # Table data
    header = ['Date', 'Session', 'USN', 'Name', 'Check In', 'Status']
    data = [header]

    for record, name, sess_date, sess_name in records:
        data.append([
            str(sess_date),
            sess_name or '',
            record.usn,
            name or '',
            str(record.check_in_time.strftime('%H:%M') if record.check_in_time else '—'),
            record.status,
        ])

    if len(data) == 1:
        data.append(['No records found for this period', '', '', '', '', ''])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#11161C')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),

        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),

        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD2D1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F4F2')]),

        # Status column coloring
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))

    elements.append(table)

    # Summary footer
    elements.append(Spacer(1, 20))
    total = len(data) - 1
    present = sum(1 for r in records if r[0].status in ('PRESENT', 'LATE'))
    summary_text = f"Total Records: {total} | Present: {present} | Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    elements.append(Paragraph(summary_text, styles['Normal']))

    doc.build(elements)
    buffer.seek(0)

    filename = f"attendance_{start}_{end}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _generate_excel(records, start, end):
    """Generate an Excel attendance report with formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styles
    header_fill = PatternFill(start_color="11161C", end_color="11161C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    present_fill = PatternFill(start_color="DCEDE9", end_color="DCEDE9", fill_type="solid")
    absent_fill = PatternFill(start_color="FFE3D9", end_color="FFE3D9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CBD2D1'),
        right=Side(style='thin', color='CBD2D1'),
        top=Side(style='thin', color='CBD2D1'),
        bottom=Side(style='thin', color='CBD2D1'),
    )

    # Title row
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Inframe Attendance Report — {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Date', 'Session', 'USN', 'Name', 'Check In', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for i, (record, name, sess_date, sess_name) in enumerate(records, 4):
        row_data = [
            str(sess_date),
            sess_name or '',
            record.usn,
            name or '',
            record.check_in_time.strftime('%H:%M') if record.check_in_time else '—',
            record.status,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Color-code status
        status_cell = ws.cell(row=i, column=6)
        if record.status in ('PRESENT', 'LATE'):
            status_cell.fill = present_fill
        elif record.status == 'ABSENT':
            status_cell.fill = absent_fill

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_length + 4, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"attendance_{start}_{end}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
